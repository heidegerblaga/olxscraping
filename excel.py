from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from main import location,price,title

wb=load_workbook("DATA.xlsx")
ws=wb.active

for row in range(1,11):
    for col in range(1,5):
         char = get_column_letter(col)
         ws[char + str(row)]= location
         print(ws[char + str(row)].value)

wb.save('DATA.xlsx')